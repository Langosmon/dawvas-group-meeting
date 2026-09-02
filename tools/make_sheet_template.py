#!/usr/bin/env python3
"""
Build the Google Sheet template (as .xlsx, to import into Google Sheets) from the
current schedule.json:

    pip install openpyxl
    python tools/make_sheet_template.py [--repo Langosmon/dawvas-group-meeting] [-o file.xlsx]

Tabs
  Availability  one row per meeting date, one column per person. Type anything in
                your cell (e.g. "out") to mark yourself unavailable that day.
                Extra columns for special sessions, cancelled meetings, overrides.
                Presenters/Notes columns look up the computed schedule from the
                Schedule tab.
  Schedule      =IMPORTDATA(...) of schedule.csv from the (public) GitHub repo, so
                the sheet always shows who actually presents.
  How to        one-screen instructions for the lab.
"""
import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import rotation as r  # noqa: E402

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
AUTO_FILL = PatternFill("solid", fgColor="EDEFF2")
OUT_FILL = PatternFill("solid", fgColor="F8C9C4")
SPECIAL_FILL = PatternFill("solid", fgColor="FFE699")
NO_MEETING_FILL = PatternFill("solid", fgColor="D9D9D9")
OVERRIDE_FILL = PatternFill("solid", fgColor="CFE2F3")
THIN = Side(style="thin", color="D0D4DA")


def build(repo: str, out: Path) -> None:
    cfg = r.load_config()
    rows = r.sheet_template_rows(cfg)
    header, data = rows[0], rows[1:]
    n_rows = len(data) + 1
    col = {name: i + 1 for i, name in enumerate(header)}
    people_first, people_last = col[cfg.rotation[0]], col[cfg.rotation[-1]]
    L = get_column_letter

    wb = Workbook()

    # ---------------------------------------------------------------- Availability
    ws = wb.active
    ws.title = "Availability"
    ws.append(header)
    for row in data:
        ws.append(row)

    # formulas for the auto columns (work whether dates are text or real dates)
    for i in range(2, n_rows + 1):
        a = f"A{i}"
        for name, idx in (("Presenters (auto)", 3), ("Notes (auto)", 4)):
            c = col[name]
            ws.cell(row=i, column=c).value = (
                f'=IFERROR(VLOOKUP({a},Schedule!$A:$D,{idx},FALSE),'
                f'IFERROR(VLOOKUP(DATEVALUE({a}),Schedule!$A:$D,{idx},FALSE),'
                f'IFERROR(VLOOKUP(TEXT({a},"yyyy-mm-dd"),Schedule!$A:$D,{idx},FALSE),"")))'
            )

    # styling
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34
    for i in range(2, n_rows + 1):
        for c in range(1, len(header) + 1):
            cell = ws.cell(row=i, column=c)
            cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
            if c <= 4:
                cell.fill = AUTO_FILL
                cell.font = Font(color="555555")
            else:
                cell.alignment = Alignment(horizontal="center")
    widths = {"Date": 12, "Day": 6, "Presenters (auto)": 22, "Notes (auto)": 34,
              "Special session (title)": 30, "Led by": 12, "No meeting (reason)": 20, "Override presenters": 20}
    for name, c in col.items():
        ws.column_dimensions[L(c)].width = widths.get(name, 11)
    ws.freeze_panes = "E2"

    rng_people = f"{L(people_first)}2:{L(people_last)}{n_rows}"
    ws.conditional_formatting.add(rng_people, FormulaRule(formula=[f"NOT(ISBLANK({L(people_first)}2))"], fill=OUT_FILL))
    for name, fill in (("Special session (title)", SPECIAL_FILL), ("No meeting (reason)", NO_MEETING_FILL),
                       ("Override presenters", OVERRIDE_FILL)):
        c = L(col[name])
        ws.conditional_formatting.add(f"{c}2:{c}{n_rows}", FormulaRule(formula=[f"NOT(ISBLANK({c}2))"], fill=fill))
    c = L(col["Led by"])
    ws.conditional_formatting.add(f"{c}2:{c}{n_rows}", FormulaRule(formula=[f"NOT(ISBLANK({c}2))"], fill=SPECIAL_FILL))

    # ---------------------------------------------------------------- Schedule
    sc = wb.create_sheet("Schedule")
    sc["A1"] = f'=IMPORTDATA("https://raw.githubusercontent.com/{repo}/main/schedule.csv")'
    sc["F1"] = "This tab pulls the computed schedule from GitHub (refreshes about hourly)."
    sc["F2"] = "If it shows an error, the repository must be public, or paste the formula from the How to tab."
    sc["F1"].font = Font(italic=True, color="555555")
    sc["F2"].font = Font(italic=True, color="555555")
    for c, w in zip("ABCD", (12, 6, 34, 44)):
        sc.column_dimensions[c].width = w

    # ---------------------------------------------------------------- How to
    ht = wb.create_sheet("How to")
    days = " & ".join(r.WEEKDAY_NAMES[d] for d in sorted(cfg.meeting_days))
    lines = [
        ("Group meeting rotation — how this sheet works", True),
        ("", False),
        (f"Meetings: {days}, {cfg.meeting_time}. Two people present per meeting, in this order: "
         f"{' → '.join(cfg.rotation)}.", False),
        ("The bot reads the Availability tab every 30 minutes, works out who presents, posts any change to "
         f"Slack ({cfg.channel}), and reminds everyone at 10:00 AM the day before each meeting.", False),
        ("", False),
        ("CAN'T PRESENT ON A DATE?", True),
        ("Find the row for that date and type anything in YOUR column — e.g. \"out\" or \"AMS conference\". "
         "Leave the cell blank if you can present.", False),
        ("The next scheduled person swaps dates with you; nobody else moves. The Presenters column shows the "
         "result within ~30 min (the Slack message is faster).", False),
        ("", False),
        ("ADVISOR / SOMEONE TAKES THE SLOT (workshop, AI training, guest)?", True),
        ("Type the title in \"Special session (title)\" and who leads it in \"Led by\". The rotation pauses that "
         "day — nobody loses their turn.", False),
        ("", False),
        ("NO MEETING AT ALL (break, travel, cancelled)?", True),
        ("Type the reason in \"No meeting (reason)\".", False),
        ("", False),
        ("NEED SPECIFIC PEOPLE ON A DATE (a swap you arranged, a make-up talk)?", True),
        ("Type their names, comma-separated, in \"Override presenters\". The people they displace get the "
         "override-people's next turns, so everyone still presents equally often.", False),
        ("", False),
        ("PLEASE DON'T", True),
        ("• edit the grey columns (Date, Day, Presenters, Notes) — they're automatic", False),
        ("• delete or reorder rows, or rename the column headers / this tab", False),
        ("• use the sheet for anything else — one row per meeting date only", False),
        ("", False),
        ("Made a mistake? Just clear the cell. Questions: ask " + cfg.organizer + ".", False),
        ("", False),
        ("Schedule tab formula, in case it needs re-pasting into Schedule!A1:", True),
        (f'=IMPORTDATA("https://raw.githubusercontent.com/{repo}/main/schedule.csv")', False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        c = ht.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=12 if bold else 11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ht.column_dimensions["A"].width = 110

    wb.save(out)
    print(f"wrote {out}  ({n_rows - 1} meeting dates, {len(cfg.rotation)} people)")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--repo", default="Langosmon/dawvas-group-meeting", help="GitHub owner/name used for the Schedule tab link")
    ap.add_argument("-o", "--out", default="DAWVAS-group-meeting-availability.xlsx")
    args = ap.parse_args()
    build(args.repo, Path(args.out))
